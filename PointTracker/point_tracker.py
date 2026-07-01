import xlsxwriter
from openpyxl.utils import get_column_letter
from datetime import datetime

POINTS_MAX_MENTOR = 60
POINTS_MAX_ATTENDANCE = 20
POINTS_MAX_STUDENTS = 20

POINTS_FOR_FIVE = 90

POINTS_MAX_TO_GIVE = 40
POINTS_COST_ABSENCE = 1

def parse_data(filename="Data.txt"):
    data = []

    with open(filename, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()

            if not line or line.startswith("#"):
                continue

            parts = line.rsplit(" ", 2)
            name = parts[0]
            dt = None
            if parts[1] != "-" and parts[2] != "-":
                dt_str = parts[1] + " " + parts[2]
                dt = datetime.strptime(dt_str, "%Y-%m-%d %H:%M")
            data.append((name, dt))

    return data, len(data)

if __name__ == "__main__":
    data, count_data = parse_data("Data.txt")

    wb = xlsxwriter.Workbook("PointTracker.xlsx")
    ws = wb.add_worksheet("Table")

    f_pres_and_slot = wb.add_format({
        "bold": True,
        "bg_color": "#ffe699",
        "align": "center",
        "border": 1
    })

    f_stud_ver = wb.add_format({
        "bg_color": "#fff2cc",
        "align": "center",
        "border": 1
    })

    f_stud_hor = wb.add_format({
        "bold": True,
        "bg_color": "#b4c7e7",
        "align": "center",
        "border": 1
    })

    f_score = wb.add_format({
        "bg_color": "#d6dce5",
        "align": "center",
        "border": 1,
        "locked": False
    })

    f_score_wrong = wb.add_format({
        "bg_color": "#ffc7ce",
        "align": "center",
        "border": 1,
        "locked": False
    })

    f_score_empty = wb.add_format({
        "bg_color": "#767171",
        "align": "center",
        "border": 1
    })

    f_agg_int = wb.add_format({
        "bg_color": "#dae3f3",
        "align": "center",
        "border": 1,
        "num_format": "0"
    })

    f_agg_dec = wb.add_format({
        "bg_color": "#dae3f3",
        "align": "center",
        "border": 1,
        "num_format": "0.00"
    })

    f_agg_mentor = wb.add_format({
        "bg_color": "#dae3f3",
        "align": "center",
        "border": 1,
        "num_format": "0",
        "locked": False
    })

    f_agg_mentor_wrong = wb.add_format({
        "bg_color": "#ffc7ce",
        "align": "center",
        "border": 1,
        "num_format": "0",
        "locked": False
    })

    f_agg_info = wb.add_format({
        "bg_color": "#ffe699",
        "align": "center",
        "border": 1
    })

    f_agg_info_bold = wb.add_format({
        "bold": True,
        "bg_color": "#ffe699",
        "align": "center",
        "border": 1
    })

    f_total_base = wb.add_format({
        "align": "center",
        "border": 1,
        "num_format": "0.00",
        "bg_color": "#ffc7ce"
    })

    f_total_high = wb.add_format({
        "bg_color": "#e2f0d9"
    })

    f_info_important_text = wb.add_format({
        "bold": True,
        "bg_color": "#f8cbad",
        "align": "center",
        "border": 1
    })

    f_info_important_data = wb.add_format({
        "bold": True,
        "bg_color": "#fbe5d6",
        "align": "center",
        "border": 1
    })

    f_info = wb.add_format({
        "align": "center",
        "valign": "vcenter",
        "text_wrap": True,
        "border": 1
    })
    
    # Predetermined header part
    ws.write(0, 0, "Presentation", f_pres_and_slot)
    ws.write(0, 1, "Slot", f_pres_and_slot)

    ws.set_column(0, 0, max(25, max(len(name) for name, _ in data)))
    ws.set_column(8 + count_data, 8 + count_data, max(len(name) for name, _ in data))
    ws.set_column(1, 1, 16)

    # Students and slots
    for i, (name, dt) in enumerate(data):
        ws.write(1 + i, 0, name, f_stud_ver) # Left names
        ws.write(1 + i, 8 + count_data, name, f_stud_ver) # Right names
        ws.write(0, 2 + i, name, f_stud_hor) # Top names
        ws.write(5 + count_data, 2 + i, name, f_stud_hor) # Bottom names
        ws.set_column(2 + i, 2 + i, len(name))
        if dt is not None: # Slots
            ws.write(1 + i, 1, dt.strftime("%Y-%m-%d %H:%M"), f_stud_ver)
        else:
            ws.write(1 + i, 1, "-", f_stud_ver)

    # Bottom part
    ws.write(2 + count_data, 1, "# attendance", f_agg_info)
    ws.write(2 + count_data, 2 + count_data, "# attendance", f_agg_info)
    ws.write(3 + count_data, 1, "P attendance", f_agg_info_bold)
    ws.write(3 + count_data, 2 + count_data, "P attendance", f_agg_info_bold)
    ws.write(4 + count_data, 1, "P total", f_agg_info)
    ws.write(4 + count_data, 2 + count_data, "P total", f_agg_info)
    ws.set_column(2 + count_data, 3 + count_data, 16)

    # Right part
    ws.write(0, 3 + count_data, "P mentor", f_agg_info_bold)
    ws.write(1 + count_data, 3 + count_data, "P mentor", f_agg_info_bold)
    ws.write(0, 4 + count_data, "sum P stud", f_agg_info)
    ws.write(1 + count_data, 4 + count_data, "sum P stud", f_agg_info)
    ws.write(0, 5 + count_data, "max P stud", f_agg_info)
    ws.write(1 + count_data, 5 + count_data, "max P stud", f_agg_info)
    ws.write(0, 6 + count_data, "P stud scaled", f_agg_info_bold)
    ws.write(1 + count_data, 6 + count_data, "P stud scaled", f_agg_info_bold)
    ws.write(0, 7 + count_data, "P total", f_agg_info)
    ws.write(1 + count_data, 7 + count_data, "P total", f_agg_info)
    ws.set_column(3 + count_data, 7 + count_data, 12)

    # Info part
    ws.write(7 + count_data, 0, "# students", f_info_important_text)
    ws.write(8 + count_data, 0, count_data, f_info_important_data)
    ws.write(10 + count_data, 0, "P for 5", f_info_important_text)
    ws.write(11 + count_data, 0, POINTS_FOR_FIVE, f_info_important_data)
    ws.write(13 + count_data, 0, f"Find your column and score the other students with whole number scores [0, {POINTS_MAX_TO_GIVE}] without repeating the same score. Use '-' if you were absent (DO NOT LEAVE EMPTY)", f_info)
    ws.write(14 + count_data, 0, f"P mentor (max {POINTS_MAX_MENTOR}) - Change if you know how many you got", f_info)
    ws.write(15 + count_data, 0, f"P attendance (max {POINTS_MAX_ATTENDANCE}) - Each absence costs 1 point", f_info)
    ws.write(16 + count_data, 0, f"P stud scaled (max {POINTS_MAX_STUDENTS}) - Ratio of obtained and maximum possible from all who were present", f_info)
    ws.set_row(13 + count_data, 90)
    ws.set_row(14 + count_data, 45)
    ws.set_row(15 + count_data, 45)
    ws.set_row(16 + count_data, 60)

    # Scores
    for i in range(count_data):
        for j in range(count_data):
            if i == j:
                ws.write(1 + i, 2 + j, "", f_score_empty)
            else:
                ws.write(1 + i, 2 + j, "-", f_score)

    ws.conditional_format(f"C2:{get_column_letter(2 + count_data)}{1 + count_data}", {
        "type": "cell",
        "criteria": "not between",
        "minimum": 0,
        "maximum": POINTS_MAX_TO_GIVE,
        "format": f_score_wrong
    })

    # Attendance points
    for i in range(count_data):
        ws.write(2 + count_data, 2 + i, f"=COUNT({get_column_letter(3 + i)}2:{get_column_letter(3 + i)}{1 + count_data})", f_agg_int)
        ws.write(3 + count_data, 2 + i, f"={POINTS_MAX_ATTENDANCE} - (A{9 + count_data} - 1 - {get_column_letter(3 + i)}{3 + count_data} * {POINTS_COST_ABSENCE})", f_agg_dec)

    # Mentor points
    for i in range(count_data):
        ws.write(1 + i, 3 + count_data, POINTS_MAX_MENTOR, f_agg_mentor)

    ws.conditional_format(f"{get_column_letter(4 + count_data)}2:{get_column_letter(4 + count_data)}{1 + count_data}", {
        "type": "cell",
        "criteria": "greater than",
        "value": f"={POINTS_MAX_MENTOR}",
        "format": f_agg_mentor_wrong
    })
    
    # Student points
    for i in range(count_data):
        ws.write(1 + i, 4 + count_data, f"=SUM(C{2 + i}:{get_column_letter(2 + count_data)}{2 + i})", f_agg_int)
        ws.write(1 + i, 5 + count_data, f"=COUNT(C{2 + i}:{get_column_letter(2 + count_data)}{2 + i}) * {POINTS_MAX_TO_GIVE}", f_agg_int)
        ws.write(1 + i, 6 + count_data, f"=({get_column_letter(5 + count_data)}{2 + i} / {get_column_letter(6 + count_data)}{2 + i}) * {POINTS_MAX_STUDENTS}", f_agg_dec)

    # Total
    for i in range(count_data):
        ws.write(4 + count_data, 2 + i, f"={get_column_letter(3 + i)}{4 + count_data} + {get_column_letter(4 + count_data)}{2 + i} + {get_column_letter(7 + count_data)}{2 + i}", f_total_base)
        ws.write(1 + i, 7 + count_data, f"={get_column_letter(3 + i)}{4 + count_data} + {get_column_letter(4 + count_data)}{2 + i} + {get_column_letter(7 + count_data)}{2 + i}", f_total_base)

    total_ranges = [
        f"C{5 + count_data}:{get_column_letter(2 + count_data)}{5 + count_data}",
        f"{get_column_letter(8 + count_data)}2:{get_column_letter(8 + count_data)}{1 + count_data}"
    ]

    for total_range in total_ranges:
        ws.conditional_format(total_range, {
            "type": "cell",
            "criteria": "greater than or equal to",
            "value": f"=$A${12 + count_data}",
            "format": f_total_high
        })

    ws.freeze_panes(1, 2)

    ws.protect()

    wb.close()